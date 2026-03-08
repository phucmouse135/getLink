"""
gui.py
======
Tkinter GUI for the Check-Mail automation tool (IMAP × GMX × Facebook/Meta).

Classes
-------
ConfigManager   – stores schema, delimiter, timeouts, and all user preferences.
WorkerThread    – runs parallel IMAP checks; communicates via a queue (thread-safe).
AutomationGUI   – builds and manages the full Tkinter interface.

Run:
    python gui.py
"""

from __future__ import annotations

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import queue
import re
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk

import config
from core import check_one_account
from utils import setup_logger

log = setup_logger("gui")


# ═══════════════════════════════════════════════════════════════
# MODULE-LEVEL CONSTANTS
# ═══════════════════════════════════════════════════════════════

APP_TITLE = "CheckMail Tool v1.0"
APP_SIZE  = "960x620"

# Fixed 4-column schema — no user configuration needed.
ALL_COLUMNS: List[str] = ["EMAIL", "PASSWORD", "STATUS", "EVIDENCE"]

_COL_WIDTH: Dict[str, int] = {
    "EMAIL":    220,
    "PASSWORD": 140,
    "STATUS":    90,
    "EVIDENCE": 460,
}

_COL_ANCHOR: Dict[str, str] = {
    "STATUS": "center",
}

_STATUS_TAG_BG: Dict[str, str] = {
    "found":     "#c8f7c5",
    "not_found": "#ffffff",
    "error":     "#ffd6d6",
    "running":   "#fff3cd",
    "pending":   "#f4f4f4",
}

_STATUS_LABEL: Dict[str, str] = {
    "found":     "✅ Found",
    "not_found": "❌ Not Found",
    "error":     "⚠️ Error",
}


# ═══════════════════════════════════════════════════════════════
# CONFIG MANAGER
# ═══════════════════════════════════════════════════════════════

class ConfigManager:
    """Runtime configuration (IMAP settings only — schema is fixed)."""

    def __init__(self) -> None:
        self.delimiter:   str = "|"
        self.subject:     str = config.TARGET_SUBJECT
        self.max_retries: int = config.MAX_RETRIES
        self.timeout:     int = config.IMAP_TIMEOUT
        self.workers:     int = config.MAX_WORKERS

    def apply(self) -> None:
        """Push current values into the global config and IMAP modules."""
        config.TARGET_SUBJECT = self.subject
        config.MAX_RETRIES    = self.max_retries
        config.IMAP_TIMEOUT   = self.timeout
        config.MAX_WORKERS    = self.workers
        try:
            import check_facebook_email as cfe  # noqa: PLC0415
            cfe.IMAP_TIMEOUT = self.timeout
            cfe.MAX_RETRIES  = self.max_retries
        except Exception:
            pass


# ═══════════════════════════════════════════════════════════════
# RESULT PARSER (module-level helper)
# ═══════════════════════════════════════════════════════════════

def _parse_result(status: str, detail: str) -> Dict[str, str]:
    """
    Convert a check_one_account result dict into Treeview column values.

    detail (found):     "UID=x | DATE=... | SUBJECT='...' | SNIPPET='...'"
    detail (not_found): "No password-change email found."
    detail (error):     "AUTH_FAIL | [E001] (...) message"
    """
    data: Dict[str, str] = {"STATUS": _STATUS_LABEL.get(status, status.upper())}

    if status == "found":
        # Parse structured segments from core.py detail string
        parts: Dict[str, str] = {}
        for seg in detail.split("|"):
            seg = seg.strip()
            for key in ("UID", "DATE", "SUBJECT", "SNIPPET"):
                if seg.startswith(f"{key}="):
                    parts[key] = seg[len(key) + 1:].strip().strip("'\"")
                    break

        subject = parts.get("SUBJECT", "")
        snippet = parts.get("SNIPPET", "")
        date    = parts.get("DATE",    "")

        # Build compact evidence: Subject first, then first 120 chars of body
        evidence_parts = []
        if subject:
            evidence_parts.append(f"[{subject}]")
        if snippet:
            evidence_parts.append(snippet[:160].replace("\n", " ").strip())
        if date and not evidence_parts:
            evidence_parts.append(date)

        data["EVIDENCE"] = "  ".join(evidence_parts)[:300]

    elif status == "error":
        msg = re.sub(r"^\[E\d+\]\s*\([^)]+\)\s*", "", detail)
        data["EVIDENCE"] = msg[:200]

    else:
        data["EVIDENCE"] = "No password-change email found."

    return data


# ═══════════════════════════════════════════════════════════════
# WORKER THREAD
# ═══════════════════════════════════════════════════════════════

class WorkerThread(threading.Thread):
    """
    Coordinate parallel IMAP checks; push all updates to the GUI via a Queue.

    Queue message shapes
    --------------------
    row_start : { "type": "row_start", "uid": str }
    row_done  : { "type": "row_done",  "uid": str, "status": str, "data": dict }
    progress  : { "type": "progress",  "done": int, "total": int,
                                       "found": int, "not_found": int, "errors": int }
    finished  : { "type": "finished" }
    """

    def __init__(
        self,
        rows: List[Tuple[str, str, str]],  # (uid, email, password)
        q:    queue.Queue,
        stop: threading.Event,
        cfg:  ConfigManager,
    ) -> None:
        super().__init__(daemon=True, name="WorkerCoordinator")
        self.rows  = rows
        self.q     = q
        self.stop  = stop
        self.cfg   = cfg

        self._lock  = threading.Lock()
        self._done  = self._found = self._nf = self._err = 0
        self._total = len(rows)

    # ── Main run ────────────────────────────────────────────────

    def run(self) -> None:
        self.cfg.apply()

        with ThreadPoolExecutor(
            max_workers=self.cfg.workers,
            thread_name_prefix="chk",
        ) as pool:
            futures: Dict[Any, str] = {}

            for uid, email, password in self.rows:
                if self.stop.is_set():
                    break
                self.q.put({"type": "row_start", "uid": uid})
                futures[pool.submit(check_one_account, email, password)] = uid

            for future in as_completed(futures):
                uid = futures[future]
                try:
                    res = future.result()
                except Exception as exc:
                    res = {"status": "error", "detail": str(exc)}
                self._emit(uid, res)

        self.q.put({"type": "finished"})

    # ── Internal helpers ─────────────────────────────────────────

    def _emit(self, uid: str, res: Dict[str, Any]) -> None:
        status = res.get("status", "error")
        detail = res.get("detail", "")
        data   = _parse_result(status, detail)

        with self._lock:
            self._done += 1
            if   status == "found":     self._found += 1
            elif status == "not_found": self._nf    += 1
            else:                       self._err   += 1
            d, t, f, n, e = (
                self._done, self._total,
                self._found, self._nf, self._err,
            )

        self.q.put({"type": "row_done",  "uid": uid, "status": status, "data": data})
        self.q.put({"type": "progress",
                    "done": d, "total": t,
                    "found": f, "not_found": n, "errors": e})


# ═══════════════════════════════════════════════════════════════
# MAIN GUI
# ═══════════════════════════════════════════════════════════════

class AutomationGUI:
    """Manages the complete Tkinter interface and all user interactions."""

    POLL_MS = 80   # queue poll interval in milliseconds

    def __init__(self, root: tk.Tk) -> None:
        self.root   = root
        self.cfg    = ConfigManager()
        self.q: queue.Queue = queue.Queue()
        self.stop   = threading.Event()
        self.worker: Optional[WorkerThread] = None

        # Internal data store: uid (str) → column-value dict
        self._rows: Dict[str, Dict[str, str]] = {}

        root.title(APP_TITLE)
        root.geometry(APP_SIZE)
        root.minsize(920, 560)
        root.configure(bg="#f0f0f0")

        self._build_style()
        self._build_ui()
        self._rebuild_tree_cols()

    # ── Style ────────────────────────────────────────────────────

    def _build_style(self) -> None:
        s = ttk.Style(self.root)
        s.theme_use("clam")
        s.configure("TFrame",   background="#f0f0f0")
        s.configure("TLabel",   background="#f0f0f0", font=("Segoe UI", 9))
        s.configure("TButton",  font=("Segoe UI", 9))
        s.configure("TSpinbox", font=("Segoe UI", 9))
        s.configure("Header.TLabel",
                    font=("Segoe UI", 9, "bold"), background="#f0f0f0")
        s.configure("Treeview",
                    font=("Segoe UI", 9), rowheight=22,
                    background="white", fieldbackground="white")
        s.configure("Treeview.Heading",
                    font=("Segoe UI", 9, "bold"),
                    background="#cfd8ea", relief="flat")
        s.map("Treeview",
              background=[("selected", "#4a90d9")],
              foreground=[("selected", "white")])

    # ── UI assembly ──────────────────────────────────────────────

    def _build_ui(self) -> None:
        top = ttk.Frame(self.root, padding=4)
        top.pack(fill="x", side="top")
        self._build_frame2(top)   # Data Source (only top frame)

        mid = ttk.Frame(self.root, padding=(4, 0, 4, 0))
        mid.pack(fill="both", expand=True)
        self._build_frame3(mid)

        bot = ttk.Frame(self.root, padding=(4, 2, 4, 4))
        bot.pack(fill="x", side="bottom")
        self._build_frame4(bot)

    # ── Frame 2: Data Source ──────────────────────────────────────

    def _build_frame2(self, parent: ttk.Frame) -> None:
        f = ttk.LabelFrame(parent, text=" 📂 Data Source ", padding=(8, 4))
        f.pack(fill="x")

        ttk.Label(f, text="Source File:", style="Header.TLabel").grid(
            row=0, column=0, sticky="w", pady=2)
        self._v_filepath = tk.StringVar()
        ttk.Entry(f, textvariable=self._v_filepath).grid(
            row=0, column=1, sticky="ew", padx=4)
        ttk.Button(f, text="Browse", command=self._browse, width=8).grid(
            row=0, column=2)

        btn_f = ttk.Frame(f)
        btn_f.grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 0))
        for text, cmd, w in [
            ("📥 Load File",   self._load_file,   12),
            ("🔄 Reset Table", self._reset_table, 12),
            ("📋 Quick Paste", self._quick_paste, 12),
        ]:
            ttk.Button(btn_f, text=text, command=cmd, width=w).pack(
                side="left", padx=(0, 4))

        self._lbl_count = ttk.Label(f, text="Rows loaded: 0", foreground="#555")
        self._lbl_count.grid(row=2, column=0, columnspan=3, sticky="w", pady=(4, 0))

        ttk.Label(
            f,
            text="Format: email|password  (one per line)",
            foreground="#888", font=("Segoe UI", 8),
        ).grid(row=3, column=0, columnspan=3, sticky="w")

        f.columnconfigure(1, weight=1)

    # ── Frame 3: Data Table ───────────────────────────────────────

    def _build_frame3(self, parent: ttk.Frame) -> None:
        f = ttk.LabelFrame(parent, text=" 📋 Data ", padding=4)
        f.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(f, selectmode="extended", show="headings")

        vsb = ttk.Scrollbar(f, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(f, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        for tag, bg in _STATUS_TAG_BG.items():
            self.tree.tag_configure(tag, background=bg)

        # Right-click context menu
        self._ctx = tk.Menu(self.root, tearoff=0)
        self._ctx.add_command(label="📋 Copy Email",       command=self._ctx_copy_email)
        self._ctx.add_command(label="📋 Copy Full Row",    command=self._ctx_copy_row)
        self._ctx.add_separator()
        self._ctx.add_command(label="🗑  Delete Selected", command=self._ctx_delete)
        self.tree.bind("<Button-3>", self._show_ctx)

    # ── Frame 4: Control & Statistics ────────────────────────────

    def _build_frame4(self, parent: ttk.Frame) -> None:
        # ── Button row ──────────────────────────────────────────
        br = ttk.Frame(parent)
        br.pack(fill="x")

        self._btn_start = ttk.Button(
            br, text="▶  START", command=self._start, width=10)
        self._btn_start.pack(side="left", padx=(0, 4))

        self._btn_stop = ttk.Button(
            br, text="■  STOP", command=self._stop, state="disabled", width=10)
        self._btn_stop.pack(side="left", padx=(0, 10))

        ttk.Button(br, text="🗑 Clear", command=self._reset_table, width=8).pack(
            side="left", padx=(0, 4))

        ttk.Button(br, text="✂ Delete Row", command=self._ctx_delete, width=11).pack(
            side="left", padx=(0, 12))

        ttk.Separator(br, orient="vertical").pack(side="left", fill="y", padx=4)

        ttk.Label(br, text="Threads:").pack(side="left", padx=(6, 2))
        self._v_workers = tk.IntVar(value=self.cfg.workers)
        ttk.Spinbox(
            br, from_=1, to=50, width=4,
            textvariable=self._v_workers,
        ).pack(side="left", padx=(0, 12))

        ttk.Separator(br, orient="vertical").pack(side="left", fill="y", padx=4)

        for label, mode in [
            ("⬇ Found",     "found"),
            ("⬇ Not Found", "not_found"),
            ("⬇ Errors",    "error"),
            ("⬇ All",       "all"),
        ]:
            ttk.Button(
                br, text=label, width=13,
                command=lambda m=mode: self._export(m),
            ).pack(side="left", padx=2)

        # ── Stats row ────────────────────────────────────────────
        sr = ttk.Frame(parent)
        sr.pack(fill="x", pady=(4, 0))

        self._v_progress = tk.StringVar(value="Progress: 0 / 0")
        self._v_found    = tk.StringVar(value="✅ Found: 0")
        self._v_nf       = tk.StringVar(value="❌ Not Found: 0")
        self._v_err      = tk.StringVar(value="⚠️ Errors: 0")
        self._v_rate     = tk.StringVar(value="Rate: 0.0%")

        _stat_labels = [
            (self._v_progress, "#333"),
            (self._v_found,    "#287028"),
            (self._v_nf,       "#555"),
            (self._v_err,      "#a02020"),
            (self._v_rate,     "#1a5fa3"),
        ]
        for var, fg in _stat_labels:
            ttk.Label(sr, textvariable=var, foreground=fg,
                      font=("Segoe UI", 9, "bold")).pack(side="left", padx=8)

        self._lbl_status = ttk.Label(
            sr, text="● Ready", foreground="#555",
            font=("Segoe UI", 9, "bold"))
        self._lbl_status.pack(side="right", padx=8)

    # ── Tree columns (fixed) ─────────────────────────────────────

    def _rebuild_tree_cols(self) -> None:
        """Set up the fixed 4 columns."""
        self.tree["columns"] = ALL_COLUMNS
        for col in ALL_COLUMNS:
            self.tree.heading(col, text=col, anchor="center")
            self.tree.column(
                col,
                width    = _COL_WIDTH.get(col, 120),
                minwidth = 50,
                anchor   = _COL_ANCHOR.get(col, "w"),
                stretch  = (col == "EVIDENCE"),
            )

    # ── File / Data operations ───────────────────────────────────

    def _browse(self) -> None:
        path = filedialog.askopenfilename(
            title="Select input file",
            filetypes=[("Text / CSV", "*.txt *.csv"), ("All files", "*.*")],
            parent=self.root,
        )
        if path:
            self._v_filepath.set(path)

    def _load_file(self) -> None:
        path = self._v_filepath.get().strip()
        if not path:
            messagebox.showwarning("No File", "Select a source file first.", parent=self.root)
            return
        if not Path(path).exists():
            messagebox.showerror("Not Found", f"File not found:\n{path}", parent=self.root)
            return

        rows = self._parse_file(path)
        if not rows:
            messagebox.showwarning("Empty", "No valid rows found in file.", parent=self.root)
            return
        self._load_rows(rows)

    def _parse_file(self, path: str) -> List[Dict[str, str]]:
        """
        Parse a .txt or .csv file into EMAIL / PASSWORD dicts.
        Accepts any of these delimiters: | : ; \t space
        """
        delimiters = ["|", ":", ";", "\t", " "]
        result: List[Dict[str, str]] = []

        if Path(path).suffix.lower() == ".csv":
            with open(path, newline="", encoding="utf-8", errors="replace") as fh:
                header_seen = False
                for parts in csv.reader(fh):
                    if not any(parts):
                        continue
                    if not header_seen and parts[0].lower() in (
                        "email", "user", "login", "account",
                    ):
                        header_seen = True
                        continue
                    header_seen = True
                    em = parts[0].strip() if len(parts) > 0 else ""
                    pw = parts[1].strip() if len(parts) > 1 else ""
                    if em:
                        result.append({"EMAIL": em, "PASSWORD": pw})
        else:
            with open(path, encoding="utf-8", errors="replace") as fh:
                for raw in fh:
                    line = raw.strip()
                    if not line or line.startswith("#"):
                        continue
                    em, pw = "", ""
                    for dlm in delimiters:
                        if dlm in line:
                            parts = line.split(dlm, maxsplit=1)
                            em, pw = parts[0].strip(), parts[1].strip()
                            break
                    else:
                        em = line  # single token — treat as email only
                    if em:
                        result.append({"EMAIL": em, "PASSWORD": pw})

        return result

    def _load_rows(self, rows: List[Dict[str, str]]) -> None:
        """Replace the table with a fresh set of rows (clears all existing data)."""
        self.tree.delete(*self.tree.get_children())
        self._rows.clear()
        self._append_rows(rows)

    def _append_rows(self, rows: List[Dict[str, str]]) -> None:
        """Append rows to the existing table without clearing it."""
        # Find next unique integer uid
        existing = {int(k) for k in self._rows if k.isdigit()}
        next_id  = max(existing, default=0) + 1

        for row_data in rows:
            while str(next_id) in self._rows:
                next_id += 1
            uid = str(next_id)
            next_id += 1
            row_data["STATUS"]   = "⏸ Pending"
            row_data["EVIDENCE"] = ""
            values = [row_data.get(c, "") for c in ALL_COLUMNS]
            self.tree.insert("", "end", iid=uid, values=values, tags=("pending",))
            self._rows[uid] = row_data

        self._lbl_count.config(text=f"Rows loaded: {len(self._rows)}")
        self._update_stats(0, len(self._rows), 0, 0, 0)

    def _quick_paste(self) -> None:
        """Open a dialog for raw text paste; parse by space and load into the table."""
        dlg = tk.Toplevel(self.root)
        dlg.title("Quick Paste")
        dlg.geometry("540x400")
        dlg.resizable(True, True)
        dlg.transient(self.root)
        dlg.grab_set()

        ttk.Label(
            dlg,
            text="Paste data — one account per line, separated by space:  email password",
            padding=(8, 6, 8, 2),
        ).pack(fill="x")

        # ── Button row (packed BEFORE the text area so it is never squeezed out) ──
        btn_f = ttk.Frame(dlg, padding=(8, 4))
        btn_f.pack(fill="x", side="bottom")

        # ── Row counter (also anchored to bottom, above buttons) ──
        lbl_rows = ttk.Label(dlg, text="Lines: 0", foreground="#555",
                             font=("Segoe UI", 8), padding=(8, 2))
        lbl_rows.pack(fill="x", side="bottom")

        # ── Text area fills the remaining space ──
        txt_f = ttk.Frame(dlg, padding=(8, 0))
        txt_f.pack(fill="both", expand=True)
        txt = tk.Text(txt_f, font=("Courier New", 9), wrap="none", undo=True)
        vsb = ttk.Scrollbar(txt_f, orient="vertical",   command=txt.yview)
        hsb = ttk.Scrollbar(txt_f, orient="horizontal", command=txt.xview)
        txt.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        txt.pack(fill="both", expand=True)

        def _update_count(*_) -> None:
            lines = [l for l in txt.get("1.0", "end").splitlines() if l.strip()]
            lbl_rows.config(text=f"Lines: {len(lines)}")
            txt.edit_modified(False)

        txt.bind("<<Modified>>", _update_count)

        def _parse_paste_text(raw: str) -> List[Dict[str, str]]:
            rows: List[Dict[str, str]] = []
            for line in raw.splitlines():
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split(None, 1)
                em = parts[0].strip()
                pw = parts[1].strip() if len(parts) > 1 else ""
                if em:
                    rows.append({"EMAIL": em, "PASSWORD": pw})
            return rows

        def _submit() -> None:
            raw = txt.get("1.0", "end").strip()
            if not raw:
                messagebox.showwarning("Empty", "Nothing to load.", parent=dlg)
                return
            rows = _parse_paste_text(raw)
            if rows:
                dlg.destroy()
                self._append_rows(rows)   # insert after existing rows
            else:
                messagebox.showwarning("Empty", "Could not parse any rows.", parent=dlg)

        def _clear() -> None:
            txt.delete("1.0", "end")
            lbl_rows.config(text="Lines: 0")

        ttk.Button(btn_f, text="✔ Submit", command=_submit,     width=10).pack(side="right", padx=4)
        ttk.Button(btn_f, text="✖ Cancel", command=dlg.destroy, width=10).pack(side="right", padx=2)
        ttk.Button(btn_f, text="🗑 Clear",  command=_clear,      width=10).pack(side="left")

    def _reset_table(self) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showwarning("Running", "Stop the task before resetting.", parent=self.root)
            return
        self.tree.delete(*self.tree.get_children())
        self._rows.clear()
        self._lbl_count.config(text="Rows loaded: 0")
        self._update_stats(0, 0, 0, 0, 0)
        self._set_status("● Ready", "#555")

    # ── Automation control ───────────────────────────────────────

    def _start(self) -> None:
        if self.worker and self.worker.is_alive():
            messagebox.showinfo("Running", "A check is already in progress.", parent=self.root)
            return

        todo: List[Tuple[str, str, str]] = []
        for uid, row in self._rows.items():
            if "✅" in row.get("STATUS", ""):
                continue   # skip already-found rows
            em = row.get("EMAIL", "").strip()
            pw = row.get("PASSWORD", "").strip()
            if em and pw:
                todo.append((uid, em, pw))

        if not todo:
            messagebox.showinfo(
                "Nothing to do",
                "No pending rows. Load a file first, or clear already-found rows.",
                parent=self.root,
            )
            return

        self.cfg.workers = max(1, self._v_workers.get())
        self.stop.clear()
        self._btn_start.config(state="disabled")
        self._btn_stop.config(state="normal", text="■  STOP")
        self._set_status("● Running", "#287028")

        self.worker = WorkerThread(todo, self.q, self.stop, self.cfg)
        self.worker.start()
        self.root.after(self.POLL_MS, self._poll)

    def _stop(self) -> None:
        if not (self.worker and self.worker.is_alive()):
            return
        self.stop.set()
        self._btn_stop.config(state="disabled", text="⏳ Stopping…")
        self._set_status("⏳ Stopping…", "#a06010")

    # ── Queue polling (Thread → GUI bridge) ──────────────────────

    def _poll(self) -> None:
        """Drain the queue and reschedule until the worker finishes."""
        try:
            while True:
                self._handle(self.q.get_nowait())
        except queue.Empty:
            pass

        if self.worker and self.worker.is_alive():
            self.root.after(self.POLL_MS, self._poll)
        else:
            # Final drain
            try:
                while True:
                    self._handle(self.q.get_nowait())
            except queue.Empty:
                pass

    def _handle(self, msg: Dict[str, Any]) -> None:
        t = msg["type"]

        if t == "row_start":
            self._set_row(msg["uid"], {"STATUS": "⏳ Running…"}, "running")

        elif t == "row_done":
            uid  = msg["uid"]
            data = msg.get("data", {})
            self._set_row(uid, data, msg["status"])
            if uid in self._rows:
                self._rows[uid].update(data)

        elif t == "progress":
            self._update_stats(
                msg["done"], msg["total"],
                msg["found"], msg["not_found"], msg["errors"],
            )

        elif t == "finished":
            self._btn_start.config(state="normal")
            self._btn_stop.config(state="disabled", text="■  STOP")
            txt   = "● Stopped" if self.stop.is_set() else "✔ Done"
            color = "#a06010"   if self.stop.is_set() else "#287028"
            self._set_status(txt, color)

    def _set_row(self, uid: str, data: Dict[str, str], tag: str) -> None:
        if not self.tree.exists(uid):
            return
        for col, val in data.items():
            try:
                self.tree.set(uid, col, val)
            except tk.TclError:
                pass   # column not in current view
        self.tree.item(uid, tags=(tag,))

    # ── Stats display ─────────────────────────────────────────────

    def _update_stats(
        self, done: int, total: int, found: int, nf: int, err: int,
    ) -> None:
        rate = (found / done * 100) if done else 0.0
        self._v_progress.set(f"Progress: {done} / {total}")
        self._v_found.set(f"✅ Found: {found}")
        self._v_nf.set(f"❌ Not Found: {nf}")
        self._v_err.set(f"⚠️ Errors: {err}")
        self._v_rate.set(f"Rate: {rate:.1f}%")

    def _set_status(self, text: str, color: str = "#555") -> None:
        self._lbl_status.config(text=text, foreground=color)

    # ── Context menu ──────────────────────────────────────────────

    def _show_ctx(self, event: tk.Event) -> None:
        iid = self.tree.identify_row(event.y)
        if iid and iid not in self.tree.selection():
            self.tree.selection_set(iid)
        try:
            self._ctx.tk_popup(event.x_root, event.y_root)
        finally:
            self._ctx.grab_release()

    def _ctx_copy_email(self) -> None:
        lines = [
            self.tree.set(iid, "EMAIL")
            for iid in self.tree.selection()
            if self.tree.exists(iid)
        ]
        self._to_clipboard(lines)

    def _ctx_copy_row(self) -> None:
        cols  = list(self.tree["columns"])
        lines = [
            "|".join(self.tree.set(iid, c) for c in cols)
            for iid in self.tree.selection()
            if self.tree.exists(iid)
        ]
        self._to_clipboard(lines)

    def _ctx_delete(self) -> None:
        sel = self.tree.selection()
        if not sel:
            return
        if not messagebox.askyesno(
            "Delete", f"Delete {len(sel)} row(s)?", parent=self.root
        ):
            return
        for iid in sel:
            self.tree.delete(iid)
            self._rows.pop(iid, None)
        self._lbl_count.config(text=f"Rows loaded: {len(self._rows)}")

    def _to_clipboard(self, lines: List[str]) -> None:
        if lines:
            self.root.clipboard_clear()
            self.root.clipboard_append("\n".join(lines))

    # ── Export ────────────────────────────────────────────────────

    def _export(self, mode: str) -> None:
        _filters = {
            "found":     lambda r: "✅" in r.get("STATUS", ""),
            "not_found": lambda r: "❌" in r.get("STATUS", ""),
            "error":     lambda r: "⚠️" in r.get("STATUS", ""),
            "all":       lambda r: True,
        }
        subset = {
            uid: row
            for uid, row in self._rows.items()
            if _filters.get(mode, lambda r: True)(row)
        }

        if not subset:
            messagebox.showinfo("Export", "No matching rows to export.", parent=self.root)
            return

        label = {"found": "Found", "not_found": "Not_Found",
                 "error": "Errors", "all": "All"}.get(mode, mode)

        if not messagebox.askyesno(
            "Confirm Export",
            f"Export {len(subset)} {label} row(s)?",
            parent=self.root,
        ):
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"export_{label.lower()}.xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("All", "*.*")],
            parent=self.root,
        )
        if not path:
            return

        try:
            self._write_xlsx(path, list(subset.values()), label)
            messagebox.showinfo(
                "Exported",
                f"{len(subset)} row(s) saved to:\n{path}",
                parent=self.root,
            )
        except Exception as exc:
            messagebox.showerror("Export Error", str(exc), parent=self.root)

    # ── Excel writer ──────────────────────────────────────────────

    def _write_xlsx(self, path: str, rows: list, sheet_name: str = "Results") -> None:
        """Write rows to a formatted .xlsx file using openpyxl."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit

        # ── Header row styling ────────────────────────────────────
        hdr_fill = PatternFill("solid", fgColor="2E75B6")
        hdr_font = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
        hdr_align = Alignment(horizontal="center", vertical="center")

        # Status-row fill colours (match GUI tags)
        _ROW_FILL = {
            "✅": PatternFill("solid", fgColor="C8F7C5"),
            "❌": PatternFill("solid", fgColor="FFFFFF"),
            "⚠️": PatternFill("solid", fgColor="FFD6D6"),
        }
        _col_widths = {
            "EMAIL":    32,
            "PASSWORD": 20,
            "STATUS":   14,
            "EVIDENCE": 70,
        }

        for ci, col in enumerate(ALL_COLUMNS, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align
            ws.column_dimensions[get_column_letter(ci)].width = _col_widths.get(col, 18)

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"  # freeze header

        # ── Data rows ─────────────────────────────────────────────
        data_font  = Font(name="Segoe UI", size=9)
        data_align = Alignment(vertical="center", wrap_text=False)

        for ri, row in enumerate(rows, start=2):
            status_val = row.get("STATUS", "")
            row_fill   = next(
                (fill for icon, fill in _ROW_FILL.items() if icon in status_val),
                None,
            )
            for ci, col in enumerate(ALL_COLUMNS, start=1):
                cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
                cell.font      = data_font
                cell.alignment = data_align
                if row_fill:
                    cell.fill = row_fill

        # Auto-filter on header row
        ws.auto_filter.ref = ws.dimensions

        wb.save(path)


# ═══════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════

def _on_close(root: tk.Tk, app: AutomationGUI) -> None:
    if app.worker and app.worker.is_alive():
        if not messagebox.askyesno(
            "Quit", "A check is in progress. Stop and quit?", parent=root
        ):
            return
        app.stop.set()
    root.destroy()


def main() -> None:
    root = tk.Tk()
    app  = AutomationGUI(root)
    root.protocol("WM_DELETE_WINDOW", lambda: _on_close(root, app))
    root.mainloop()


if __name__ == "__main__":
    main()
